import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from CoordsMessages import CoordinatesTrackedMessage


class CoordsWriter:
    def __init__(self, filepath: str):
        self.path = filepath
        self.workbook: Workbook = openpyxl.load_workbook(self.path)
        self.sheet: Worksheet = self.workbook.active

    def __del__(self):
        # pass
        self.workbook.save(self.path)

    def write(self, msg: CoordinatesTrackedMessage):
        row = (int)(msg.t / 0.5 + 2)

        self.sheet.cell(row=row, column=1).value = msg.t
        self.sheet.cell(row=row, column=2).value = msg.x
        self.sheet.cell(row=row, column=3).value = msg.y
        self.sheet.cell(row=row, column=4).value = msg.z


# test
if __name__ == '__main__':
    path = "data/videoset0/Seq0_settings.xlsx"

    writer = CoordsWriter(path)

    msg1 = CoordinatesTrackedMessage(2.5, 1.7, -1.3, 2.5, 1)
    msg2 = CoordinatesTrackedMessage(2.5, 1.7, -2.3, 3.0, 1)

    writer.write(msg1)
    writer.write(msg2)

    del writer

