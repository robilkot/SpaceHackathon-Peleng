import threading

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from Models.Camera import Camera
from Models.ObjectState import ObjectState
from Constants import *


class ExcelWriter:
    def __init__(self, filepath: str):
        self.lock = threading.Lock()
        self.path = filepath
        self.workbook: Workbook = openpyxl.load_workbook(self.path)
        self.sheet: Worksheet = self.workbook.active

    def write(self, msg: ObjectState):
        row = int(msg.t / 0.5 + 2)
        column_offset = 0  # todo: shifts output to another column for debug

        # self.sheet_out.cell(row=row, column=1+column_offset).value = msg.t
        self.sheet.cell(row=row, column=2+column_offset).value = msg.x
        self.sheet.cell(row=row, column=3+column_offset).value = msg.y
        self.sheet.cell(row=row, column=4+column_offset).value = msg.z
        self.sheet.cell(row=row, column=5+column_offset).value = msg.dl_max

        if DEBUG_WRITER:
            print(f'write t:{msg.t} x:{msg.x} y:{msg.y} z:{msg.z}')

        self.lock.acquire()
        self.workbook.save(self.path)
        self.lock.release()

    def read_params(self) -> dict[int, Camera]:
        col = 8

        focal_length = self.sheet.cell(row=3, column=col).value
        matrix_width = self.sheet.cell(row=4, column=col).value
        matrix_height = self.sheet.cell(row=5, column=col).value

        cameras = {}
        for i in range(CAMERAS_COUNT):
            offset = i * 7
            x = self.sheet.cell(row=9 + offset, column=col).value
            y = self.sheet.cell(row=10 + offset, column=col).value
            z = self.sheet.cell(row=11 + offset, column=col).value
            a = self.sheet.cell(row=12 + offset, column=col).value

            camera = Camera(focal_length,
                            x, y, z, a,
                            matrix_width,
                            matrix_height,
                            RESOLUTION[0],
                            RESOLUTION[1])
            cameras[i+1] = camera

            print(f"read camera {camera}")

        if DEBUG_WRITER:
            print(f'read {len(cameras)} cameras setup')

        return cameras


# test
if __name__ == '__main__':
    path = "data/videoset0/Seq0_settings.xlsx"

    writer = ExcelWriter(path)

    msg1 = ObjectState(2.5, 1.7, -1.3, 2.5, None, None, None)
    msg2 = ObjectState(2.5, 1.7, -2.3, 3.0, None, None, None)

    writer.write(msg1)
    writer.write(msg2)

    del writer
